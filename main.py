# External libraries:
import os
import shutil
import datetime
import openpyxl

# Local class objects:
from teacher import Teacher
from lp import LessonPlan

# Folder variables:
LP_UOI_TAG: str = 'UOI'
LP_ESL_TAG: str = 'English' 
LP_FOLDER_PATH: str = r'N:\2022-2023上学期\International Classes国际班'
LP_UOI_FOLDER_PATH: str = os.path.join(LP_FOLDER_PATH, LP_UOI_TAG)
LP_ESL_FOLDER_PATH: str = os.path.join(LP_FOLDER_PATH, LP_ESL_TAG)


def connect_network():

    # Address and user settings:
    NETWORK_ADDRESS: str = r'\\KindergartenNAS\HC-Kindergarden'
    NETWORK_DRIVE_ASSIGNED: str = 'N:'
    USER_REQUIRED: bool = False
    USER_NAME: str = 'teacher'
    USER_PASSWORD: str = 'hc@laoshi123'
    
    # Establishing new connection:
    command_prompt: str = f'net use {NETWORK_DRIVE_ASSIGNED} {NETWORK_ADDRESS} /p:yes '
    if USER_REQUIRED: command_prompt += '/user:{USER_NAME} {USER_PASSWORD}'
    os.system(command_prompt)


def get_teachers_list():

    # Obtaining teacher and class information:
    teachers_list = []
    for (dirname, dirnames, filenames) in os.walk(LP_UOI_FOLDER_PATH):

        # Class-specific folders:
        for class_id in dirnames:
            class_id_folder_path: str = os.path.join(LP_UOI_FOLDER_PATH, class_id)
            for (dirname, dirnames, filenames) in os.walk(class_id_folder_path):
                
                # Getting teacher name and setting up Teacher class object:
                for teacher_name in dirnames:
                    
                    # General information:
                    teacher = Teacher()
                    teacher.name: str = teacher_name
                    teacher.class_id: str = class_id

                    # Teacher-specific lesson plan folders:
                    uoi_folder_path: str = os.path.join(class_id_folder_path, teacher_name)
                    esl_folder_path: str = uoi_folder_path.replace(LP_UOI_TAG, LP_ESL_TAG)
                    teacher.lp_uoi_folder_path: str = uoi_folder_path
                    teacher.lp_esl_folder_path: str = esl_folder_path
                    teachers_list.append(teacher)
                break
        break

    # Returning Teacher class objects list:
    return teachers_list


def get_lesson_plans_list(path: str):
    
    # Obtaining lesson plans from a path:
    lesson_plans_list: list = []
    for (dirname, dirnames, filenames) in os.walk(path):
        if len(filenames) > 0:
            for filename in filenames:

                # Creating Lesson Plan class object:
                lesson_plan = LessonPlan()
                lesson_plan.filename_source: str = filename
                filepath_source = os.path.join(dirname, filename)
                lesson_plan.filepath_source: str = filepath_source

                # Adding lesson plan class object to the list:
                lesson_plans_list.append(lesson_plan)
    
    # Returning Lesson Plan class objects list:
    return lesson_plans_list


def get_timestamp():
    
    # Getting datetime value:
    dt_now = datetime.datetime.now()

    # Reformatting datetime value to YYYYMMDD_HHMMSS format:
    dt_timestamp_format: str = '%y%m%d_%H%M%S'
    dt_timestamp: str = datetime.datetime.strftime(dt_now, dt_timestamp_format)
    
    # Returning timestamp value:
    return dt_timestamp


def make_session_folder_map(teachers_list: list, session_timestamp: str):
    
    # Getting class names list:
    class_id_list: list = []
    for teacher in teachers_list:
        teacher: Teacher
        if teacher.class_id not in class_id_list:
            class_id_list.append(teacher.class_id)
    
    # Creating and mapping session folder:
    project_folder_path: str = os.path.dirname(__file__)
    session_time_stamp: str = session_timestamp
    session_folder_name: str = session_time_stamp
    session_folder_path: str = os.path.join(project_folder_path, session_folder_name)
    os.mkdir(session_folder_path)

    # Creating class folders:
    # ...\ICKA-1
    # ...\ICKA-2
    # >>
    for class_id in class_id_list:
        class_id_folder_path: str = os.path.join(session_folder_path, class_id)
        os.mkdir(class_id_folder_path)

        # Creating teachers folders:
        # ...\ICKC-2\Andy
        # ...\ICKC-2\Elsie
        # ...\ICKC-2\Wing
        # >>
        for teacher in teachers_list:
            teacher: Teacher
            if teacher.class_id == class_id:
                teacher_name_folder_path: str = os.path.join(class_id_folder_path, teacher.name)
                os.mkdir(teacher_name_folder_path)

                # Creating subject specific folders:
                # ...\ICKC-2\Andy\UOI
                # ...\ICKC-2\Andy\English
                subject_tags = (LP_UOI_TAG, LP_ESL_TAG)
                for subject_tag in subject_tags:
                    subject_tag: str
                    subject_tag_folder_path: str = os.path.join(teacher_name_folder_path, subject_tag)
                    os.mkdir(subject_tag_folder_path)


def write_results(path: str, teachers_list: list):

    # Workbook attributes:
    WORKBOOK_FILENAME: str = 'results.xlsx'
    WORKBOOK_FILEPATH: str = os.path.join(path, WORKBOOK_FILENAME)

    # Worksheet attributes:
    WEEK_RANGE: int = 24
    RESERVED_ROW: int = 1
    HEADER_CLASS_NAME_COLUMN: str = 'A'
    HEADER_TEACHER_NAME_COLUMN: str = 'B'
    HEADER_WEEK_DICT: dict = {}

    # Creating new workbook:
    workbook = openpyxl.Workbook()

    # Creating worksheets:
    subject_tags: list = []
    for teacher in teachers_list:
        teacher: Teacher
        for subject_tag in teacher.lp_list:
            subject_tag: str
            subject_tags.append(subject_tag)
            if subject_tag not in workbook.sheetnames:
                workbook.create_sheet(subject_tag)
    default_worksheet_name: str = 'Sheet'
    del workbook[default_worksheet_name]
    
    # Preparing workbook by adding basic structure:
    for subject_tag in subject_tags:
        worksheet = workbook[subject_tag]

        # Writing headers:
        cell_header_class_name: str = f'{HEADER_CLASS_NAME_COLUMN}{RESERVED_ROW}'
        worksheet[cell_header_class_name] = 'Class'
        cell_header_teacher_name: str = f'{HEADER_TEACHER_NAME_COLUMN}{RESERVED_ROW}'
        worksheet[cell_header_teacher_name] = 'Teacher'

        # Writing weeks:
        alphabet: str = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        alphabet_index: int = alphabet.index(HEADER_TEACHER_NAME_COLUMN) + 1
        for week in range(1, WEEK_RANGE + 1):
            row: int = RESERVED_ROW
            column: str = alphabet[alphabet_index]
            cell_header_week: str = f'{column}{row}'
            worksheet[cell_header_week] = week
            HEADER_WEEK_DICT[week] = column
            alphabet_index += 1

        # Writing results:
        row = RESERVED_ROW + 1
        for teacher in teachers_list:
            teacher: Teacher

            # Class and teacher information:
            cell_class_id: str = f'{HEADER_CLASS_NAME_COLUMN}{row}'
            worksheet[cell_class_id] = teacher.class_id
            cell_teacher_name: str = f'{HEADER_TEACHER_NAME_COLUMN}{row}'
            worksheet[cell_teacher_name] = teacher.name

            # Lesson plans:
            for lesson_plan in teacher.lp_list[subject_tag]:
                lesson_plan: LessonPlan
                if lesson_plan.week_num != 0 :
                    column: int = HEADER_WEEK_DICT[lesson_plan.week_num]
                    cell_lesson_plan: str = f'{column}{row}'
                    worksheet[cell_lesson_plan] = '+'
            
            # Next teacher:
            row += 1

    # Saving workbook:
    workbook.save(filename=WORKBOOK_FILEPATH)


def main():

    # Checking and establishing connection:
    if not os.path.isdir(LP_FOLDER_PATH): connect_network()

    # Obtaining teachers list and lesson plans:
    teachers_list: list = get_teachers_list()
    for teacher in teachers_list:
        teacher: Teacher
        teacher.lp_list[LP_UOI_TAG] = get_lesson_plans_list(teacher.lp_uoi_folder_path)
        teacher.lp_list[LP_ESL_TAG] = get_lesson_plans_list(teacher.lp_esl_folder_path)

    # Mapping new session folder:
    project_folder_path: str = os.path.dirname(__file__)
    session_timestamp: str = get_timestamp()
    make_session_folder_map(teachers_list, session_timestamp)

    # Working with lesson plans:
    for teacher in teachers_list:
        teacher: Teacher
        subject_tags = (LP_UOI_TAG, LP_ESL_TAG)
        for subject_tag in subject_tags:
            for lesson_plan in teacher.lp_list[subject_tag]:
                lesson_plan: LessonPlan

                # Updating lesson plan's week number:
                lesson_plan.update_week_num()

                # Locating correct folder:
                session_folder: str = os.path.join(project_folder_path, session_timestamp)
                class_id_folder_save: str = os.path.join(session_folder, teacher.class_id)
                teacher_name_folder_save: str = os.path.join(class_id_folder_save, teacher.name)
                subject_tag_folder_save: str = os.path.join(teacher_name_folder_save, subject_tag)

                # Generating new filename and filepath for copy:
                filename_copy: str = f'{lesson_plan.week_num}.docx'
                filepath_copy: str = os.path.join(subject_tag_folder_save, filename_copy)
                lesson_plan.filename_copy: str = filename_copy
                lesson_plan.filepath_copy: str = filepath_copy

                # Making a copy:
                if lesson_plan.week_num > 0:
                    shutil.copy2(lesson_plan.filepath_source, lesson_plan.filepath_copy)
                else:
                    log_message: str = f'{teacher.name}\'s \"{lesson_plan.filename_source}\" has no week, file was not copied'
                    print(log_message)
    
    # Writing results:
    write_results(session_folder, teachers_list)


if __name__ == '__main__':
    main()
